#Bibliotecas usadas
from openpyxl import Workbook
import os
from time import sleep
from rich.console import Console, Group
from rich.panel import Panel
from rich.align import Align
from rich.prompt import Prompt

console = Console()

def menu():
    titulo = Align.center("[bold cyan]AUTOPY[/bold cyan]")
    instrucoes = Align.center("[bold yellow]Bot conversor de TXT para Excel[/bold yellow]")
    painel_interno = Panel(Group(titulo, instrucoes), border_style="cyan", width=50)

    texto_externo = Align.center("[bold magenta]Bem-vindo ao programa![/bold magenta]")
    creditos = Align.center("""
[bold magenta]
Feito por: Lucas
           Ranny
           Gerson
           Felipe
        [/bold magenta]""")
    
    painel_interno_centro = Align.center(painel_interno)
    conteudo_externo = Group(texto_externo, painel_interno_centro, creditos)
    painel_externo = Panel(conteudo_externo, border_style="magenta", width=70)

    console.print("\n")
    console.print(painel_externo)
    console.print("\n")

def solicita_arquivo_txt():
    menu()
    while True:    
        url = Prompt.ask("Digite a url do arquivo de texto (.txt)").strip()

        if not url:
            console.print("[red]Programa cancelado![/red]")
            exit(0)

        if os.path.isfile(url) and url.lower().endswith(".txt"):
            console.print("[green]Arquivo localizado com sucesso![/green]")
            sleep(1)
            return url
        console.print("[red]url inválido/arquivo não é .txt. Por favor, tente novamente[/red]")
    
def nome_excel():
    while True:
        os.system('cls')
        menu()
        nome = Prompt.ask("Digite o nome da planilha a ser salva (não precisa escrever .xlxl):")
        if not nome:
            console.print("[red]Nome inválido. Tente novamente.[/red]")
            continue
        if not nome.lower().endswith(".xlsx"):
            nome += ".xlsx"
        console.print("[green]Gerando Planilha...[/green]")
        sleep(2)
        return nome

def extrair_dados(linhas):
    registros = []
    registro = {}
    for linha in linhas:
        linha = linha.strip()
        if not linha:
            if registro:
                registros.append(registro)
                registro = {}
            continue
        if ":" in linha:
            chave, valor = linha.split(":", 1)
            registro[chave.strip()] = valor.strip()
    if registro:
        registros.append(registro)
    return registros

def gerar_planilha(dados, nome_arquivo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"

    cabecalhos = list(dados[0].keys())
    for col, cabecalho in enumerate(cabecalhos,start=1):
        ws.cell(row=1, column=col, value=cabecalho)

     # Dados
    for linha_idx, registro in enumerate(dados, start=2):
        for col_idx, cabecalho in enumerate(cabecalhos, start=1):
            ws.cell(row=linha_idx, column=col_idx, value=registro.get(cabecalho, ""))
    wb.save(nome_arquivo)

def main():
    arquivo = solicita_arquivo_txt()
    with open(arquivo, "r", encoding="utf-8") as arquivo:
        linhas = arquivo.readlines()
    dados = extrair_dados(linhas)

    if dados:
        arquivo_excel = nome_excel()
        gerar_planilha(dados,arquivo_excel)
        console.print('[green]Planilha criada com sucesso![/green]')
        console.print("[cyan]Planilha está salva na pasta do protótipo[/cyan]")
    else:
        console.print("[red]Nada válido encontrado.[/red]")
main()